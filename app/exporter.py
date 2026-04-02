"""Excel workbook exporter — produces a real .xlsx with 4 sheets."""

from __future__ import annotations

import logging
from collections import Counter
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.models import Lead, SourceLog

logger = logging.getLogger(__name__)

# Header style
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Lead_Inbox column order (exact spec)
LEAD_INBOX_COLUMNS: list[str] = [
    "lead_id",
    "collected_at",
    "company_name",
    "lead_lane",
    "portfolio_type",
    "private_company_confirmed",
    "state",
    "city",
    "website",
    "business_phone",
    "email",
    "named_contact",
    "contact_title",
    "employee_estimate",
    "distress_signal",
    "financing_signal",
    "bankruptcy_chapter",
    "reason_qualified",
    "quality_tier",
    "confidence_score",
    "source_type",
    "source_url",
    "notes",
    "status",
]

SOURCES_LOG_COLUMNS: list[str] = [
    "source_name",
    "source_type",
    "source_url",
    "collected_at",
    "leads_found",
    "leads_kept",
    "notes",
]


def _style_header(ws, col_count: int) -> None:
    """Apply header styling to row 1."""
    for col_idx in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN


def _auto_width(ws, col_count: int, max_width: int = 40) -> None:
    """Auto-size columns based on content."""
    for col_idx in range(1, col_count + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=False):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, max_width)


def _write_summary(ws, leads: list[Lead]) -> None:
    """Write the Summary sheet."""
    ws.append(["Lead Pipeline — Export Summary"])
    ws.merge_cells("A1:B1")
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])

    ws.append(["Export Timestamp", datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")])
    ws.append(["Total Leads Kept", len(leads)])
    ws.append([])

    # By lane
    ws.append(["Leads by Lane"])
    ws[f"A{ws.max_row}"].font = Font(bold=True, size=12)
    lane_counts = Counter(ld.lead_lane.value for ld in leads)
    for lane, count in sorted(lane_counts.items()):
        ws.append(["", lane, count])

    ws.append([])

    # By quality tier
    ws.append(["Leads by Quality Tier"])
    ws[f"A{ws.max_row}"].font = Font(bold=True, size=12)
    tier_counts = Counter(ld.quality_tier.value for ld in leads)
    for tier, count in sorted(tier_counts.items()):
        ws.append(["", tier, count])

    ws.append([])

    # By state
    ws.append(["Leads by State"])
    ws[f"A{ws.max_row}"].font = Font(bold=True, size=12)
    state_counts = Counter(ld.state for ld in leads)
    for state, count in sorted(state_counts.items()):
        ws.append(["", state, count])

    ws.append([])

    # By source
    ws.append(["Leads by Source"])
    ws[f"A{ws.max_row}"].font = Font(bold=True, size=12)
    source_counts = Counter(ld.source_type.value for ld in leads)
    for src, count in sorted(source_counts.items()):
        ws.append(["", src, count])

    _auto_width(ws, 3)


def _write_lead_inbox(ws, leads: list[Lead]) -> None:
    """Write the Lead_Inbox sheet with exact column order."""
    ws.append(LEAD_INBOX_COLUMNS)
    _style_header(ws, len(LEAD_INBOX_COLUMNS))

    for lead in leads:
        data = lead.model_dump()
        # Resolve enums to their string values
        row = []
        for col in LEAD_INBOX_COLUMNS:
            val = data.get(col, "")
            if hasattr(val, "value"):
                val = val.value
            if val is None:
                val = ""
            row.append(val)
        ws.append(row)

    _auto_width(ws, len(LEAD_INBOX_COLUMNS))


def _write_rules(ws) -> None:
    """Write the Rules sheet documenting pipeline logic."""
    ws.append(["Rule Category", "Rule", "Details"])
    _style_header(ws, 3)

    rules_data = [
        ("Lane Logic", "charged_off", "Private companies selling or likely selling charged-off debt"),
        ("Lane Logic", "bankruptcy", "Chapter 13 bankruptcy leads — private companies only"),
        ("Lane Logic", "performing", "Private companies with performing portfolios"),
        ("Lane Logic", "capital_seeking", "Private companies seeking capital, lines of credit, bridge funding, etc."),
        ("", "", ""),
        ("Excluded States", "charged_off", "TX, NC, SC, PA, AZ, CA"),
        ("Excluded States", "bankruptcy", "(none)"),
        ("Excluded States", "performing", "(none)"),
        ("Excluded States", "capital_seeking", "HI, AK"),
        ("", "", ""),
        ("Quality — Best Case", "Required", "company_name, website, business_phone, reason_qualified, named_contact, contact_title"),
        ("Quality — Mid Level", "Required", "company_name, website, business_phone, reason_qualified"),
        ("Quality — Weak", "Definition", "Missing business_phone, website, or reason_qualified"),
        ("", "", ""),
        ("Discard", "weak_quality", "Discard if quality_tier == weak"),
        ("Discard", "excluded_state", "Discard if state is excluded for the lane"),
        ("Discard", "public_company", "Discard if public_company_confirmed == True"),
        ("Discard", "trustee", "Discard if trustee_related == True"),
        ("", "", ""),
        ("Preference", "employee_size", "Prefer 10–50 employees; score favorably, no hard-fail"),
        ("Preference", "ranking", "best_case ranked above mid_level; then by confidence_score desc"),
    ]
    for row in rules_data:
        ws.append(list(row))

    _auto_width(ws, 3)


def _write_sources_log(ws, source_logs: list[SourceLog]) -> None:
    """Write the Sources_Log sheet."""
    ws.append(SOURCES_LOG_COLUMNS)
    _style_header(ws, len(SOURCES_LOG_COLUMNS))

    for slog in source_logs:
        data = slog.model_dump()
        row = []
        for col in SOURCES_LOG_COLUMNS:
            val = data.get(col, "")
            if hasattr(val, "value"):
                val = val.value
            if val is None:
                val = ""
            row.append(val)
        ws.append(row)

    _auto_width(ws, len(SOURCES_LOG_COLUMNS))


def export_workbook(
    leads: list[Lead],
    source_logs: list[SourceLog],
    output_path: Path,
) -> Path:
    """Generate the full .xlsx workbook and save to output_path."""
    wb = Workbook()

    # Sheet 1: Summary (rename default sheet)
    ws_summary = wb.active
    ws_summary.title = "Summary"
    _write_summary(ws_summary, leads)

    # Sheet 2: Lead_Inbox
    ws_inbox = wb.create_sheet("Lead_Inbox")
    _write_lead_inbox(ws_inbox, leads)

    # Sheet 3: Rules
    ws_rules = wb.create_sheet("Rules")
    _write_rules(ws_rules)

    # Sheet 4: Sources_Log
    ws_sources = wb.create_sheet("Sources_Log")
    _write_sources_log(ws_sources, source_logs)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    logger.info("Workbook saved: %s", output_path)
    return output_path
